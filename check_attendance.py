"""
Author: Dr Renju Mathew

Distribution according to BSD 3-Clause License 

This script processes attendance data by comparing a list of expected participants against a list of actual participants.
It loads two CSV files: one containing the expected participants with their official and informal names, and the other containing the actual participants.
The script then replaces the informal names with the official names.
Next, it determines who attended the event, who did not attend, and who is unknown.
Finally, it creates a DataFrame with the attendance information, including the names of the attendees and their attendance status (Present, Absent, or Unknown).

Version 2. Creates an Excel file with the attendance information in a table that includes the names of the attendees, their attendance status, and their coach's name.
"""

import pandas as pd
import glob

# Load expected participants

# Search for file path that starts with "expected_participants"
file_path = glob.glob('expected_participants*.csv')[0]
expected_participants = pd.read_csv(file_path)
expected_participants_list = set(expected_participants['Official Name'])

# Change all names to lower case in expected_participants_list
expected_participants['Name (Original Name)'] = expected_participants['Name (Original Name)'].str.lower()

expected_participants_dict = dict(zip(expected_participants['Name (Original Name)'], 
                                      expected_participants['Official Name']))

# Load actual participants. Remove unnecessary rows and columns.
actual_participants_file = glob.glob('participants_*.csv')[0]  # Get the name of the csv file that begins with "participant"
df = pd.read_csv(actual_participants_file, header=None)

# Find the row where the column header is 'Name (Original Name)'
locations = df.where(df == 'Name (Original Name)').stack().index.tolist()

# Delete all rows before the first location
df = df.drop(df.index[:locations[0][0]])

# Delete all columns except the first one
df = df.drop(df.columns[1:], axis=1)

# Make the first row the header
df.columns = df.iloc[0]
df = df.iloc[1:].reset_index(drop=True)

# Set all names to lowercase
df['Name (Original Name)'] = df['Name (Original Name)'].str.lower()

# Replace full stops with spaces
df['Name (Original Name)'] = df['Name (Original Name)'].str.replace('.', ' ', regex=True)

# Replace the informal name with the official name or keep the informal name if not in the dictionary
actual_participants_list = df['Name (Original Name)'].apply(lambda x: expected_participants_dict.get(x, x)).tolist()

# Determine who attended, who did not attend, and who is unknown
present = set(actual_participants_list).intersection(expected_participants_list)
absent = set(expected_participants_list).difference(actual_participants_list)
unknown = [name.title() for name in actual_participants_list if name not in expected_participants_list]

# Create a dataframe for attendance status
attendance_df = (pd.concat([pd.DataFrame(present, columns=['Name']).assign(Status='Present'),
                            pd.DataFrame(absent, columns=['Name']).assign(Status='Absent')])
                 .sort_values('Name'))

# Add unknown attendees to the dataframe
unknown_df = pd.DataFrame(unknown, columns=['Name']).assign(Status='Present (Unrecognized Name)').reset_index(drop=True)
attendance_df = pd.concat([attendance_df, unknown_df]).reset_index(drop=True)
attendance_df.index += 1

# Create a column with the name Status Code. 1 for present. 0 for absent.
# attendance_df['Status Code'] = attendance_df['Status'].apply(lambda x: 1 if x == 'Present' else 0)

# Reorder the columns
#attendance_df = attendance_df[['Name', 'Status Code', 'Status']]
#attendance_df = attendance_df[['Name', 'Status Code', 'Status']]

attendance_df.to_csv('attendance.csv', index=False)
print(attendance_df)
print()

print("-----------------------")
try:
    expected_participants = pd.read_csv('expected_participants.csv')
    expected_participants = expected_participants[['Official Name', 'Coach Name']]
    # Drop duplicates
    expected_participants = expected_participants.drop_duplicates()
    joined_df = pd.merge(attendance_df, expected_participants, how='left', left_on='Name', right_on='Official Name').drop('Official Name', axis=1)

    # Save as an excel file, but make the data a table
    joined_df.to_excel('attendance.xlsx', index=False, sheet_name='Attendance')

    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.styles import PatternFill

    # Load the workbook and select the worksheet
    wb = load_workbook('attendance.xlsx')
    ws = wb['Attendance']

    # Create a table for the data
    table = Table(displayName="Table1", ref=ws.dimensions)

    # Add a default style to the table with striped rows
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Highlight cells with the word "Absent" in yellow
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Absent":
                cell.fill = yellow_fill
            elif cell.value == "Present (Unrecognized Name)":
                cell.fill = orange_fill

    # Save the modified workbook
    wb.save('attendance.xlsx')

    print("Saved to attendance.xlsx")
    # Remove the csv file
    import os
    os.remove('attendance.csv')
except:
    print("Saved to attendance.csv")